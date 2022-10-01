from openpyxl import workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill

#FF3C33-red colour
#61FF33 - green colour

wb = Workbook()
#create the workbook
#give the Sheet Name
wb['Sheet'].title = "Automation Report"
#active the cell -workbook
sh1 = wb.active

sh1['A1'].value = "Name"
sh1['B1'].value = "Status"
sh1['A2'].value = "Python"
sh1['B2'].value = "Active"
#lets fill the Active cell with Green colour
sh1['B2'].fill = PatternFill("solid",fgColor="257A1E")

sh1['A3'].value = "Java"
sh1['B3'].value = "Inactive"

sh1['B3'].fill = PatternFill("solid",fgColor="7A281E")

wb.save("FinalReport_New.xlsx")

