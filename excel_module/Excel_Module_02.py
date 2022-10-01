import openpyxl

#load the workbook
wb = openpyxl.load_workbook("F:\\Python\\Selenium-Python\\TestData\\Python_Practice.xlsx")

sh1 = wb['Names']

#to get the number of rows and columns
rowcount = sh1.max_row
colCount = sh1.max_column

#to print the number of rows and columns
print('Number of rows',rowcount)
print('Number of columns', colCount)

#Iterate over each row and column

for i in range(1,rowcount + 1):
    for j in range (1, colCount + 1):
        print(sh1.cell(i ,j).value)

#to write the data in the particular cell
sh1.cell(row = 4 , column = 1 , value= 'Jiten')
#to write one more record , row remains same
sh1.cell(row = 4 , column = 2 , value= 'Mumbai')
sh1.cell(row = 4 , column = 3 , value= 35)

#save the workbok in a new workbook
wb.save("F:\\Python\\Selenium-Python\\TestData\\Python_Practice_02.xlsx")