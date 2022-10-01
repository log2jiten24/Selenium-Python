import openpyxl

#load the workbook
wb = openpyxl.load_workbook("F:\\Python\\Selenium-Python\\TestData\\Python_Practice.xlsx")
#it will store inside the workbook and workbook is a class
print(type(wb))
#to get the number of sheets
sheets = wb.sheetnames
print(sheets)

#to know the active sheet in the excel where it is pointing
print(wb.active.title)

#get the sheet name which we want to work
sh1 = wb['Names']
print(type(sh1))

#to get the value of particular cell
data = sh1['A2'].value
print(data)

#Alternative way of doing it
data1 = wb['Names']['B2'].value
print(data1)

print('************************************************************')
#option 2 of retrieve the value from the excel
print(sh1.cell(3,2).value)
print(sh1.cell(3,3).value)

#Now get the value from the second sheet
sh2 = wb['Marks']
print(sh2.cell(2,1).value)
print('************************************************************')
#Another way of doing it by passing the row number and column value

print(sh2.cell(row=2 , column= 2).value)